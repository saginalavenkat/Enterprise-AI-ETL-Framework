"""
===============================================================================
File Name   : document_loader.py
Project     : Enterprise AI ETL Framework
Purpose     : Load ETL Project Documents
Author      : Venkata
===============================================================================
"""

from pathlib import Path

from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

from core.logger.logger import logger


class DocumentLoader:
    """
    Loads supported project documents and returns their text.
    """

    def __init__(self):

        self.base_dir = Path(__file__).resolve().parent.parent

        self.documents_dir = self.base_dir / "documents"

    # ------------------------------------------------------------------

    def load_pdf(self, file_name):

        logger.info(f"Loading PDF : {file_name}")

        pdf_path = self.documents_dir / file_name

        reader = PdfReader(pdf_path)

        text = ""

        for page in reader.pages:
            page_text = page.extract_text()

            if page_text:
                text += page_text + "\n"

        return text

    # ------------------------------------------------------------------

    def load_docx(self, file_name):

        logger.info(f"Loading DOCX : {file_name}")

        doc_path = self.documents_dir / file_name

        document = Document(doc_path)

        text = "\n".join(
            paragraph.text
            for paragraph in document.paragraphs
        )

        return text

    # ------------------------------------------------------------------

    def load_excel(self, file_name):

        logger.info(f"Loading Excel : {file_name}")

        excel_path = self.documents_dir / file_name

        workbook = load_workbook(excel_path)

        text = ""

        for sheet in workbook.sheetnames:

            worksheet = workbook[sheet]

            text += f"\nSheet : {sheet}\n"

            for row in worksheet.iter_rows(values_only=True):

                row_data = [str(cell) if cell is not None else "" for cell in row]

                text += " | ".join(row_data) + "\n"

        return text

    # ------------------------------------------------------------------

    def load_txt(self, file_name):

        logger.info(f"Loading TXT : {file_name}")

        txt_path = self.documents_dir / file_name

        with open(txt_path, "r", encoding="utf-8") as file:

            return file.read()

if __name__ == "__main__":
    loader = DocumentLoader()
    sttm = loader.load_excel("STTM.xlsx")
    print(sttm[:1000])

    rules = loader.load_docx("Business_Rules.docx")
    print(rules[:1000])

    design = loader.load_pdf("Design_Document.pdf")
    print(design[:1000])